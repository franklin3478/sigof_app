import streamlit as st
import pandas as pd
import requests
from bs4 import BeautifulSoup
from openpyxl import load_workbook
from io import BytesIO
from urllib.parse import urljoin
import re
import asyncio
import sys
from playwright.sync_api import sync_playwright
from reportlab.lib import colors
from reportlab.lib.pagesizes import A3, landscape
from reportlab.platypus import (SimpleDocTemplate,Table,TableStyle,Image as RLImage,PageBreak,Paragraph)
from reportlab.lib.units import cm
from reportlab.lib.utils import ImageReader
from PIL import Image as PILImage
from reportlab.lib.styles import ParagraphStyle
from reportlab.lib.enums import TA_CENTER, TA_LEFT
from concurrent.futures import ThreadPoolExecutor, as_completed


# ============================================================
# CONFIGURACIÓN
# ============================================================

TIMEOUT = 20

EXTENSIONES_IMAGEN = (
    ".jpg",
    ".jpeg",
    ".png",
    ".webp",
    ".gif"
)


# ============================================================
# CONFIGURACIÓN PLAYWRIGHT PARA WINDOWS + STREAMLIT
# ============================================================

if sys.platform == "win32":
    try:
        asyncio.set_event_loop_policy(
            asyncio.WindowsProactorEventLoopPolicy()
        )
    except Exception:
        pass

# ============================================================
# LEER EXCEL Y OBTENER HIPERVÍNCULOS
# ============================================================

@st.cache_data(show_spinner=False)
def leer_excel_con_links(archivo):

    contenido = archivo.getvalue()

    # --------------------------------------------------------
    # Leer Excel con Pandas
    # --------------------------------------------------------

    df = pd.read_excel(
        BytesIO(contenido)
    )

    df.columns = [
        str(col).strip()
        for col in df.columns
    ]

    # --------------------------------------------------------
    # Buscar columna FOTO
    # --------------------------------------------------------

    columna_foto = None

    for col in df.columns:

        if str(col).strip().lower() == "foto":

            columna_foto = col
            break

    if columna_foto is None:

        raise ValueError(
            "El Excel debe contener una columna llamada 'foto'."
        )

    # --------------------------------------------------------
    # Abrir Excel con OpenPyXL
    # --------------------------------------------------------

    wb = load_workbook(
        BytesIO(contenido),
        data_only=False
    )

    ws = wb.active

    # --------------------------------------------------------
    # Encontrar columna FOTO
    # --------------------------------------------------------

    numero_columna_foto = None

    for celda in ws[1]:

        if (
            str(celda.value)
            .strip()
            .lower()
            == "foto"
        ):

            numero_columna_foto = celda.column
            break

    if numero_columna_foto is None:

        raise ValueError(
            "No se encontró la columna 'foto'."
        )

    # --------------------------------------------------------
    # Obtener hipervínculos
    # --------------------------------------------------------

    enlaces = []

    for fila in range(
        2,
        ws.max_row + 1
    ):

        celda = ws.cell(
            row=fila,
            column=numero_columna_foto
        )

        url = None

        # ----------------------------------------------------
        # Hipervínculo real de Excel
        # ----------------------------------------------------

        if celda.hyperlink:

            url = celda.hyperlink.target

        # ----------------------------------------------------
        # Fórmula HYPERLINK()
        # ----------------------------------------------------

        elif isinstance(celda.value, str):

            valor = celda.value.strip()

            coincidencia = re.search(
                r'HYPERLINK\s*\(\s*"([^"]+)"',
                valor,
                flags=re.IGNORECASE
            )

            if coincidencia:

                url = coincidencia.group(1)

        enlaces.append(url)

    # --------------------------------------------------------
    # Ajustar cantidad de enlaces
    # --------------------------------------------------------

    if len(enlaces) < len(df):

        enlaces.extend(
            [None] * (
                len(df) -
                len(enlaces)
            )
        )

    df["__url_foto"] = enlaces[:len(df)]

    return df, columna_foto


# ============================================================
# EXTRAER ENLACES DE IMÁGENES
# ============================================================

@st.cache_data(show_spinner=False)
def extraer_imagenes(url):

    """
    Obtiene las URLs de las fotografías desde SIGOF.
    Optimizada para responder más rápido.
    NO se utiliza para FieldService.
    """

    if not url:
        return []

    url = str(url).strip()

    # ========================================================
    # FIELD SERVICE
    # NO TOCAR SU PROCESO
    # ========================================================

    if "servicios.distriluz.com.pe/FieldService" in url:
        return []

    # ========================================================
    # SIGOF
    # ========================================================

    try:

        respuesta = requests.get(
            url,
            timeout=TIMEOUT,
            verify=False,
            headers={
                "User-Agent": (
                    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                    "AppleWebKit/537.36 (KHTML, like Gecko) "
                    "Chrome/131.0.0.0 Safari/537.36"
                ),
                "Accept": "text/html,application/xhtml+xml"
            }
        )

        if respuesta.status_code != 200:
            return []

        # Si SIGOF devuelve directamente una imagen

        content_type = respuesta.headers.get(
            "Content-Type",
            ""
        ).lower()

        if "image/" in content_type:
            return [url]

        # ====================================================
        # ANALIZAR HTML
        # ====================================================

        soup = BeautifulSoup(
            respuesta.content,
            "html.parser"
        )

        imagenes = []
        vistas = set()

        # ----------------------------------------------------
        # 1. BUSCAR IMÁGENES DIRECTAMENTE
        # ----------------------------------------------------

        for img in soup.find_all("img"):

            for atributo in (
                "src",
                "data-src",
                "data-original",
                "data-lazy-src"
            ):

                valor = img.get(atributo)

                if not valor:
                    continue

                valor = str(valor).strip()

                if not valor:
                    continue

                imagen_url = urljoin(
                    url,
                    valor
                )

                if imagen_url in vistas:
                    continue

                extension = (
                    imagen_url
                    .lower()
                    .split("?")[0]
                )

                if extension.endswith(
                    EXTENSIONES_IMAGEN
                ):

                    vistas.add(imagen_url)
                    imagenes.append(imagen_url)

        # ----------------------------------------------------
        # 2. BUSCAR ENLACES A FOTOGRAFÍAS
        # ----------------------------------------------------

        for enlace in soup.find_all(
            "a",
            href=True
        ):

            href = str(
                enlace.get("href")
            ).strip()

            if not href:
                continue

            imagen_url = urljoin(
                url,
                href
            )

            if imagen_url in vistas:
                continue

            extension = (
                imagen_url
                .lower()
                .split("?")[0]
            )

            if extension.endswith(
                EXTENSIONES_IMAGEN
            ):

                vistas.add(imagen_url)
                imagenes.append(imagen_url)

        return imagenes

    except requests.RequestException:
        return []

    except Exception:
        return []


def extraer_imagenes_sigof_paralelo(
    urls,
    trabajadores=10
):

    """
    Consulta varias páginas SIGOF simultáneamente.
    La función devuelve los resultados conforme terminan,
    sin obligar a esperar el orden original.
    """

    resultados = {}
    urls_validas = []

    for url in urls:

        if not url:
            continue

        url = str(url).strip()

        if not url:
            continue

        if (
            "servicios.distriluz.com.pe/FieldService"
            in url
        ):
            continue

        urls_validas.append(url)

    if not urls_validas:
        return resultados

    urls_validas = list(
        dict.fromkeys(urls_validas)
    )

    with ThreadPoolExecutor(
        max_workers=trabajadores
    ) as executor:

        futuros = {
            executor.submit(
                extraer_imagenes,
                url
            ): url
            for url in urls_validas
        }

        for futuro in as_completed(futuros):

            url = futuros[futuro]

            try:

                resultados[url] = (
                    futuro.result()
                )

            except Exception:

                resultados[url] = []

    return resultados


# ============================================================
# DESCARGAR FOTOS FIELDSERVICE
# ============================================================
@st.cache_data(show_spinner=False)
def descargar_fotos_fieldservice(url):

    fotos = []

    try:

        with sync_playwright() as p:

            browser = p.chromium.launch(
                headless=True
            )

            context = browser.new_context(
                viewport={
                    "width": 1400,
                    "height": 1000
                },
                user_agent=(
                    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                    "AppleWebKit/537.36 (KHTML, like Gecko) "
                    "Chrome/139.0.0.0 Safari/537.36"
                )
            )

            page = context.new_page()

            # =================================================
            # ABRIR FIELDSERVICE
            # =================================================

            page.goto(
                url,
                wait_until="domcontentloaded",
                timeout=60000
            )

            # =================================================
            # ESPERAR A QUE BLAZOR CARGUE LA GALERÍA
            # =================================================

            page.wait_for_selector(
                "section.public-photo-gallery",
                timeout=30000
            )

            # =================================================
            # ESPERAR A QUE SE COMPLETE EL RENDERIZADO
            # =================================================

            page.wait_for_timeout(2000)

            # =================================================
            # OBTENER DIRECTAMENTE LOS HREF DE LAS FOTOS
            # =================================================

            urls_fotos = page.locator(
                "section.public-photo-gallery a"
            ).evaluate_all(
                """
                elementos => elementos
                    .map(e => e.href)
                    .filter(Boolean)
                """
            )

            # =================================================
            # ELIMINAR DUPLICADOS
            # =================================================

            urls_fotos = list(
                dict.fromkeys(urls_fotos)
            )

            # =================================================
            # SI NO HAY HREF, BUSCAR LOS SRC DE IMG
            # =================================================

            if not urls_fotos:

                urls_fotos = page.locator(
                    "section.public-photo-gallery img"
                ).evaluate_all(
                    """
                    elementos => elementos
                        .map(e => e.src)
                        .filter(Boolean)
                    """
                )

                urls_fotos = list(
                    dict.fromkeys(urls_fotos)
                )

            # =================================================
            # DESCARGAR DIRECTAMENTE LAS FOTOS
            # =================================================

            for url_foto in urls_fotos[:2]:

                try:

                    respuesta = context.request.get(
                        url_foto,
                        timeout=30000,
                        headers={
                            "Referer": page.url,
                            "Accept": (
                                "image/avif,image/webp,image/apng,"
                                "image/svg+xml,image/*,*/*;q=0.8"
                            )
                        }
                    )

                    if not respuesta.ok:
                        continue

                    contenido = respuesta.body()

                    if not contenido:
                        continue

                    # =================================================
                    # COMPROBAR QUE SEA UNA IMAGEN
                    # =================================================

                    try:

                        imagen = PILImage.open(
                            BytesIO(contenido)
                        )

                        imagen.verify()

                        fotos.append(
                            contenido
                        )

                    except Exception:

                        continue

                except Exception:

                    continue

            browser.close()

            return fotos[:2]

    except Exception:

        return []

# ============================================================
# DESCARGAR UNA IMAGEN DESDE URL
# ============================================================

@st.cache_data(show_spinner=False)
def descargar_imagen_url(url):

    if not url:
        return None

    try:

        respuesta = requests.get(
            url,
            headers={
                "User-Agent": (
                    "Mozilla/5.0 "
                    "(Windows NT 10.0; Win64; x64) "
                    "AppleWebKit/537.36 "
                    "(KHTML, like Gecko) "
                    "Chrome/139.0 Safari/537.36"
                )
            },
            timeout=30,
            allow_redirects=True
        )

        respuesta.raise_for_status()

        if not respuesta.content:
            return None

        return respuesta.content

    except Exception:
        return None


# ============================================================
# MOSTRAR INFORMACIÓN
# ============================================================

def mostrar_datos(
    fila,
    columnas_mostrar
):

    for columna in columnas_mostrar:

        valor = fila.get(
            columna,
            ""
        )

        # ----------------------------------------------------
        # CELDAS VACÍAS
        # ----------------------------------------------------

        if pd.isna(valor):

            valor = ""

        # ----------------------------------------------------
        # FORMATEAR NÚMEROS
        # ----------------------------------------------------

        elif isinstance(valor, float):

            if valor.is_integer():

                valor = str(
                    int(valor)
                )

            else:

                valor = str(valor)

        # ----------------------------------------------------
        # OTROS VALORES
        # ----------------------------------------------------

        else:

            valor = str(valor).strip()

        nombre = (
            str(columna)
            .replace("_", " ")
            .title()
        )

        st.markdown(
            f"**{nombre}:** {valor}"
        )


# ============================================================
# MOSTRAR FOTOS
# ============================================================

def mostrar_fotos(
    url,
    imagenes_sigof=None
):

    if not url:

        st.warning(
            "⚠️ Este registro no tiene hipervínculo."
        )

        return

    url = str(url).strip()

    # ========================================================
    # FIELDSERVICE
    # NO SE MODIFICA
    # ========================================================

    if "servicios.distriluz.com.pe/FieldService" in url:

        with st.spinner(
            "📷 Cargando fotografías..."
        ):

            fotos = descargar_fotos_fieldservice(
                url
            )

        if not fotos:

            st.warning(
                "⚠️ No se pudieron cargar las fotografías."
            )

            st.link_button(
                "🔗 Abrir enlace original",
                url
            )

            return

        st.caption(
            f"📷 {len(fotos)} fotografía(s)"
        )

        columnas = st.columns(
            min(len(fotos), 2)
        )

        for posicion, foto in enumerate(fotos):

            with columnas[
                posicion % 2
            ]:

                st.image(
                    foto,
                    use_container_width=True
                )

        return

    # ========================================================
    # SIGOF
    # ========================================================

    if imagenes_sigof is not None:

        imagenes = imagenes_sigof

    else:

        with st.spinner(
            "📷 Buscando fotografías..."
        ):

            imagenes = extraer_imagenes(
                url
            )

    if not imagenes:

        st.warning(
            "⚠️ No se encontraron imágenes "
            "en este enlace."
        )

        st.link_button(
            "🔗 Abrir enlace original",
            url
        )

        return

    st.caption(
        f"📷 {len(imagenes)} fotografía(s)"
    )

    if len(imagenes) == 1:

        st.image(
            imagenes[0],
            use_container_width=True
        )

    else:

        columnas = st.columns(
            min(len(imagenes), 2)
        )

        for posicion, imagen in enumerate(imagenes):

            with columnas[
                posicion % 2
            ]:

                st.image(
                    imagen,
                    use_container_width=True
                )


# ============================================================
# MOSTRAR UN REGISTRO
# ============================================================

def mostrar_registro(
    indice,
    fila,
    columnas_mostrar,
    imagenes_sigof=None
):

    url = fila.get(
        "__url_foto"
    )

    if pd.isna(url) or not str(url).strip():
        return

    # ========================================================
    # MARCO DEL REGISTRO
    # ========================================================

    with st.container(
        border=True
    ):

        st.markdown(
            f"### 📷 Registro {indice}"
        )

        st.divider()

        # ====================================================
        # FOTOS
        # ====================================================

        mostrar_fotos(
            url,
            imagenes_sigof=imagenes_sigof
        )

        # ====================================================
        # INFORMACIÓN
        # ====================================================

        if columnas_mostrar:

            st.divider()

            mostrar_datos(
                fila,
                columnas_mostrar
            )

        # ====================================================
        # ENLACE ORIGINAL
        # ====================================================

        with st.expander(
            "🔗 Ver enlace original"
        ):

            st.code(
                str(url),
                language="text"
            )

def mostrar_registro_progresivo(indice, fila, columnas_mostrar):
    """
    Muestra el registro inmediatamente y deja un espacio para
    cargar posteriormente la fotografía SIGOF.
    """

    url = fila.get("__url_foto")

    if isinstance(url, pd.Series):
        url = url.iloc[0] if not url.empty else ""

    if pd.isna(url):
        url = ""

    url = str(url).strip()

    es_fieldservice = "servicios.distriluz.com.pe/FieldService" in url

    with st.container(border=True):

        # ========================================================
        # ENCABEZADO
        # ========================================================
        st.markdown(f"### 📷 Registro {indice}")
        st.divider()

        # ========================================================
        # ESPACIO PARA FOTOGRAFÍA
        # ========================================================
        placeholder_foto = st.empty()

        if es_fieldservice:
            # FIELD SERVICE SE MANTIENE EXACTAMENTE IGUAL
            with placeholder_foto.container():
                mostrar_fotos(
                    url,
                    imagenes_sigof=None
                )

            placeholder_foto_resultado = None

        else:
            # SIGOF: primero mostramos el registro y el espacio
            # para que la fotografía llegue posteriormente.
            placeholder_foto.info("⏳ Cargando fotografía...")

            placeholder_foto_resultado = placeholder_foto

        # ========================================================
        # DATOS DEL REGISTRO
        # ========================================================
        for columna in columnas_mostrar:
            valor = fila.get(columna, "")

            if pd.isna(valor):
                valor = ""

            st.markdown(
                f"**{columna}:** {valor}"
            )

        # ========================================================
        # ENLACE ORIGINAL
        # ========================================================
        if url:
            with st.expander("🔗 Ver foto original"):
                st.link_button(
                    "Abrir fotografía original",
                    url,
                    use_container_width=True
                )

    return placeholder_foto_resultado

# ============================================================
# FUNCIÓN PRINCIPAL
# ============================================================

def ejecutar_galeria_lectura():

    st.title(
        "📷 Galería de Fotos Lectura"
    )

    st.caption(
        "Las fotografías se obtienen automáticamente "
        "desde el hipervínculo de la columna 'foto'."
    )

    # ========================================================
    # CARGAR EXCEL
    # ========================================================

    archivo = st.file_uploader(
        "📂 Seleccionar archivo Excel",
        type=["xlsx", "xls"],
        key="galeria_lectura_excel"
    )

    if archivo is None:

        st.info(
            "Seleccione un archivo Excel para comenzar."
        )

        return

    # ========================================================
    # IDENTIFICAR EXCEL
    # ========================================================

    archivo_id = (
        archivo.name,
        len(archivo.getvalue())
    )

    # ========================================================
    # LEER EXCEL SOLO CUANDO CAMBIA
    # ========================================================

    if (
        st.session_state.get(
            "galeria_archivo_id"
        )
        != archivo_id
    ):

        try:

            with st.spinner(
                "📂 Procesando archivo Excel..."
            ):

                df, columna_foto = leer_excel_con_links(
                    archivo
                )

            # ------------------------------------------------
            # Guardar resultados en session_state
            # ------------------------------------------------

            st.session_state.galeria_archivo_id = archivo_id

            st.session_state.galeria_df = df

            st.session_state.galeria_columna_foto = (
                columna_foto
            )

            # ------------------------------------------------
            # Reiniciar fotografías mostradas
            # ------------------------------------------------

            st.session_state.galeria_fotos_hasta = 0

            # ------------------------------------------------
            # Reiniciar resultado de filtros
            # ------------------------------------------------

            st.session_state.galeria_resultado_filtro_id = None

            # ------------------------------------------------
            # Reiniciar filtros seleccionados
            # ------------------------------------------------

            st.session_state.pop(
                "galeria_columnas_mostrar",
                None
            )

            st.session_state.pop(
                "galeria_filtros_habilitados",
                None
            )

        except Exception as e:

            st.error(
                f"❌ Error al leer el Excel:\n\n{e}"
            )

            return

    # ========================================================
    # RECUPERAR DATAFRAME
    # ========================================================

    df = st.session_state.get(
        "galeria_df"
    )

    columna_foto = st.session_state.get(
        "galeria_columna_foto"
    )

    if df is None:

        st.error(
            "❌ No se pudo cargar el archivo."
        )

        return

    # ========================================================
    # VALIDAR
    # ========================================================

    if df.empty:

        st.warning(
            "El Excel no contiene registros."
        )

        return

    # ========================================================
    # ESTADÍSTICAS
    # ========================================================

    total = len(df)

    con_link = (
        df["__url_foto"]
        .notna()
        .sum()
    )

    col1, col2 = st.columns(2)

    with col1:

        st.metric(
            "📋 Registros",
            total
        )

    with col2:

        st.metric(
            "🔗 Registros con foto",
            con_link
        )

    st.divider()

    # ========================================================
    # SIDEBAR
    # ========================================================

    st.sidebar.header(
        "⚙️ Configuración"
    )

    # ========================================================
    # COLUMNAS DISPONIBLES
    # ========================================================

    columnas_disponibles = [
        columna
        for columna in df.columns
        if columna not in [
            columna_foto,
            "__url_foto"
        ]
    ]

    # ========================================================
    # INFORMACIÓN
    # ========================================================

    st.sidebar.subheader(
        "👁️ Información"
    )

    columnas_mostrar = st.sidebar.multiselect(
        "Seleccionar columnas",
        options=columnas_disponibles,
        default=[],
        key="galeria_columnas_mostrar"
    )

    # ========================================================
    # FILTROS
    # ========================================================

    st.sidebar.subheader(
        "🔎 Filtros"
    )

    filtros_habilitados = st.sidebar.multiselect(
        "Seleccionar filtros",
        options=columnas_disponibles,
        default=[],
        help=(
            "Seleccione las columnas que desea "
            "utilizar como filtros."
        ),
        key="galeria_filtros_habilitados"
    )

    # ========================================================
    # FILTRAR REGISTROS CON FOTO
    # ========================================================

    df_filtrado = df.copy()

    df_filtrado = df_filtrado[
        df_filtrado["__url_foto"].notna()
        &
        (
            df_filtrado["__url_foto"]
            .astype(str)
            .str.strip()
            .ne("")
        )
    ].copy()

    # ========================================================
    # APLICAR FILTROS
    # ========================================================

    for columna in filtros_habilitados:

        valores_filtro = df_filtrado[
            columna
        ].copy()

        valores_filtro = valores_filtro.apply(
            lambda x: (
                str(int(x))
                if pd.notna(x)
                and isinstance(x, (int, float))
                and float(x).is_integer()
                else str(x).strip()
                if pd.notna(x)
                else "[VACÍO]"
            )
        )

        valores = sorted(
            valores_filtro.unique()
        )

        seleccion = st.sidebar.multiselect(
            f"🔎 {columna}",
            options=valores,
            key=f"filtro_valor_{columna}"
        )

        if seleccion:

            df_filtrado = df_filtrado[
                valores_filtro.isin(seleccion)
            ].copy()

    # ========================================================
    # RESULTADOS
    # ========================================================

    total_filtrado = len(
        df_filtrado
    )

    st.subheader(
        f"📸 Registros filtrados: {total_filtrado}"
    )

    if df_filtrado.empty:

        st.warning(
            "No existen registros con "
            "los filtros seleccionados."
        )

        return

    # ========================================================
    # DETECTAR CAMBIO DE RESULTADO DEL FILTRO
    # ========================================================

    resultado_filtro_id = tuple(
        df_filtrado.index.tolist()
    )

    if st.session_state.get(
        "galeria_resultado_filtro_id"
    ) != resultado_filtro_id:

        st.session_state.galeria_resultado_filtro_id = (
            resultado_filtro_id
        )

        st.session_state.galeria_fotos_hasta = 0

    # ========================================================
    # TAMAÑO DEL BLOQUE
    # ========================================================

    TAMANO_BLOQUE = 200

    fotos_hasta = st.session_state.get(
        "galeria_fotos_hasta",
        0
    )

    # ========================================================
    # TODAVÍA NO CARGAMOS FOTOGRAFÍAS
    # ========================================================

    if fotos_hasta == 0:

        st.info(
            f"📋 Hay {total_filtrado:,} registros "
            "que cumplen los filtros.\n\n"
            f"Se mostrarán en bloques de "
            f"{TAMANO_BLOQUE} registros."
        )

        if st.button(
            "🚀 Cargar fotografías",
            type="primary",
            use_container_width=True,
            key="btn_cargar_fotografias"
        ):

            siguiente = min(
                TAMANO_BLOQUE,
                total_filtrado
            )

            st.session_state.galeria_fotos_hasta = siguiente

            st.rerun()

        st.stop()

    # ========================================================
    # TOMAR SOLO LOS REGISTROS DEL BLOQUE
    # ========================================================

    registros_a_mostrar = df_filtrado.iloc[
        :fotos_hasta
    ]

    cantidad_mostrada = len(
        registros_a_mostrar
    )

    # ========================================================
    # INFORMACIÓN
    # ========================================================

    st.success(
        f"📷 Mostrando {cantidad_mostrada:,} "
        f"de {total_filtrado:,} registros"
    )

    progreso = (
        cantidad_mostrada /
        total_filtrado
    )

    st.progress(
        progreso,
        text=(
            f"{cantidad_mostrada:,} "
            f"de {total_filtrado:,} registros"
        )
    )

    st.divider()

    # ========================================================
    # MOSTRAR REGISTROS
    # ========================================================

    registros = list(
        registros_a_mostrar.iterrows()
    )

    # ============================================================
    # PREPARAR CONSULTAS SIGOF EN PARALELO
    # ============================================================

    urls_sigof = []

    for _, fila in registros:

        url = fila.get("__url_foto")

        if isinstance(url, pd.Series):
            url = url.iloc[0] if not url.empty else ""

        if pd.isna(url):
            continue

        url = str(url).strip()

        if not url:
            continue

        # Solo SIGOF
        if "servicios.distriluz.com.pe/FieldService" not in url:
            urls_sigof.append(url)

    # Eliminar URLs repetidas
    urls_sigof = list(dict.fromkeys(urls_sigof))


    # ============================================================
    # CREAR CONSULTAS SIGOF
    # ============================================================

    futuros_sigof = {}

    executor = ThreadPoolExecutor(max_workers=10)

    for url in urls_sigof:

        futuro = executor.submit(
            extraer_imagenes,
            url
        )

        futuros_sigof[futuro] = url


    # ============================================================
    # MOSTRAR GALERÍA INMEDIATAMENTE
    # ============================================================

    placeholders_sigof = {}


    for posicion in range(0, len(registros), 2):

        columnas = st.columns(2)

        # --------------------------------------------------------
        # REGISTRO IZQUIERDO
        # --------------------------------------------------------

        with columnas[0]:

            indice_real, fila = registros[posicion]

            placeholder = mostrar_registro_progresivo(
                posicion + 1,
                fila,
                columnas_mostrar
            )

            if placeholder is not None:

                url = fila.get("__url_foto")

                if isinstance(url, pd.Series):
                    url = url.iloc[0] if not url.empty else ""

                if pd.notna(url):

                    url = str(url).strip()

                    if url:
                        placeholders_sigof.setdefault(
                            url,
                            []
                        ).append(placeholder)


        # --------------------------------------------------------
        # REGISTRO DERECHO
        # --------------------------------------------------------

        if posicion + 1 < len(registros):

            with columnas[1]:

                indice_real, fila = registros[posicion + 1]

                placeholder = mostrar_registro_progresivo(
                    posicion + 2,
                    fila,
                    columnas_mostrar
                )

                if placeholder is not None:

                    url = fila.get("__url_foto")

                    if isinstance(url, pd.Series):
                        url = url.iloc[0] if not url.empty else ""

                    if pd.notna(url):

                        url = str(url).strip()

                        if url:
                            placeholders_sigof.setdefault(
                                url,
                                []
                            ).append(placeholder)


    # ============================================================
    # ACTUALIZAR LAS FOTOGRAFÍAS A MEDIDA QUE TERMINAN
    # ============================================================

    try:

        for futuro in as_completed(futuros_sigof):

            url = futuros_sigof[futuro]

            try:

                imagenes = futuro.result()

            except Exception:

                imagenes = []

            # Buscar todos los registros que usan esta misma URL
            placeholders = placeholders_sigof.get(
                url,
                []
            )

            for placeholder in placeholders:

                placeholder.empty()

                with placeholder.container():

                    if not imagenes:

                        st.warning(
                            "⚠️ No se encontraron fotografías."
                        )

                        st.link_button(
                            "🔗 Abrir fotografía original",
                            url,
                            use_container_width=True
                        )

                    else:

                        st.caption(
                            f"📸 {len(imagenes)} fotografía(s)"
                        )

                        if len(imagenes) == 1:

                            st.image(
                                imagenes[0],
                                use_container_width=True
                            )

                        else:

                            columnas_foto = st.columns(
                                min(len(imagenes), 2)
                            )

                            for i, imagen in enumerate(imagenes):

                                with columnas_foto[
                                    i % 2
                                ]:

                                    st.image(
                                        imagen,
                                        use_container_width=True
                                    )

    finally:
        executor.shutdown(wait=True)

    # ========================================================
    # SIGUIENTE BLOQUE
    # ========================================================

    if fotos_hasta < total_filtrado:

        restantes = (
            total_filtrado -
            fotos_hasta
        )

        siguiente = min(
            TAMANO_BLOQUE,
            restantes
        )

        st.divider()

        st.info(
            f"📦 Ya se cargaron {fotos_hasta:,} "
            f"registros. "
            f"Quedan {restantes:,}."
        )

        if st.button(
            f"🚀 Cargar siguientes {siguiente} registros",
            type="primary",
            use_container_width=True,
            key=f"btn_siguiente_{fotos_hasta}"
        ):

            st.session_state.galeria_fotos_hasta = (
                fotos_hasta + siguiente
            )

            st.rerun()

    else:

        st.divider()

        st.success(
            f"✅ Se han mostrado los "
            f"{total_filtrado:,} registros filtrados."
        )

    # ========================================================
    # EXPORTAR PDF CON FOTOS
    # ========================================================

    st.subheader(
        "📄 Exportar registros"
    )

    st.info(
        "El PDF se generará en formato A3 horizontal, "
        "con un registro por página y hasta dos fotografías "
        "por registro."
    )

    if st.button(
        "📥 Generar PDF con Fotos",
        type="primary",
        use_container_width=True,
        key="btn_generar_pdf_fotos"
    ):

        with st.spinner(
            "📄 Generando PDF..."
        ):

            try:

                pdf = generar_pdf_con_fotos(
                    df_filtrado
                )

                st.download_button(
                    label="⬇️ Descargar PDF",
                    data=pdf,
                    file_name="galeria_lectura.pdf",
                    mime="application/pdf",
                    use_container_width=True,
                    key="btn_descargar_pdf"
                )

                st.success(
                    "✅ PDF generado correctamente."
                )

            except Exception as e:

                st.error(
                    f"❌ Error al generar el PDF:\n\n{e}"
                )


# ============================================================
# DETECTAR SI LA URL ES FIELDSERVICE
# ============================================================

def es_fieldservice(url):

    if not url:
        return False

    return (
        "servicios.distriluz.com.pe/FieldService"
        in str(url)
    )


# ============================================================
# GENERAR PDF A3 HORIZONTAL
# 1 REGISTRO POR PÁGINA
# MÁXIMO 2 FOTOS POR REGISTRO
# ============================================================

def generar_pdf_con_fotos(df):

    archivo_salida = BytesIO()

    # ========================================================
    # A3 HORIZONTAL
    # ========================================================

    ancho_pagina, alto_pagina = landscape(A3)

    margen = 0.8 * cm

    doc = SimpleDocTemplate(
        archivo_salida,
        pagesize=landscape(A3),
        leftMargin=margen,
        rightMargin=margen,
        topMargin=margen,
        bottomMargin=margen
    )

    elementos = []

    # ========================================================
    # ESTILOS
    # ========================================================

    estilo_encabezado = ParagraphStyle(
        "EncabezadoPDF",
        fontName="Helvetica-Bold",
        fontSize=8,
        leading=9,
        alignment=TA_CENTER
    )

    estilo_dato = ParagraphStyle(
        "DatoPDF",
        fontName="Helvetica",
        fontSize=8,
        leading=9,
        alignment=TA_CENTER,
        wordWrap="CJK"
    )

    # ========================================================
    # COLUMNAS
    # ========================================================

    df = df.copy()

    columnas_normalizadas = {}

    for col in df.columns:

        nombre = re.sub(
            r"\s+",
            " ",
            str(col).strip().lower()
        )

        columnas_normalizadas[nombre] = col

    def buscar_columna(*nombres):

        for nombre in nombres:

            nombre_normalizado = re.sub(
                r"\s+",
                " ",
                nombre.strip().lower()
            )

            if (
                nombre_normalizado
                in columnas_normalizadas
            ):

                return columnas_normalizadas[
                    nombre_normalizado
                ]

        return None

    col_suministro = buscar_columna(
        "suministro",
        "nro suministro",
        "n° suministro",
        "numero suministro",
        "número suministro"
    )

    col_medidor = buscar_columna(
        "medidor",
        "nro medidor",
        "n° medidor",
        "numero medidor",
        "número medidor"
    )

    col_direccion = buscar_columna(
        "direccion",
        "dirección"
    )

    col_obs = buscar_columna(
        "obs"
    )

    col_obs_descripcion = buscar_columna(
        "obs_descripcion",
        "obs descripcion",
        "obs descripción",
        "observacion descripcion",
        "observación descripción"
    )

    col_lectura = buscar_columna(
        "lectura"
    )

    # ========================================================
    # OBTENER VALOR
    # ========================================================

    def obtener_valor(
        fila,
        columna
    ):

        if not columna:
            return ""

        valor = fila.get(
            columna,
            ""
        )

        if pd.isna(valor):
            return ""

        if isinstance(valor, float):

            if valor.is_integer():

                return str(
                    int(valor)
                )

            return str(valor)

        return str(valor).strip()

    # ========================================================
    # DESCARGAR UNA IMAGEN DEL SIGOF ANTIGUO
    # ========================================================

    def descargar_imagen_para_pdf(url):

        if not url:
            return None

        try:

            respuesta = requests.get(
                url,
                headers={
                    "User-Agent": (
                        "Mozilla/5.0 "
                        "(Windows NT 10.0; Win64; x64) "
                        "AppleWebKit/537.36 "
                        "(KHTML, like Gecko) "
                        "Chrome/139.0 Safari/537.36"
                    )
                },
                timeout=30,
                allow_redirects=True
            )

            respuesta.raise_for_status()

            contenido = respuesta.content

            if not contenido:
                return None

            try:

                imagen = PILImage.open(
                    BytesIO(contenido)
                )

                imagen.load()

                return contenido

            except Exception:

                return None

        except Exception:

            return None

    # ========================================================
    # OBTENER LAS FOTOS DEL REGISTRO
    # ========================================================

    def obtener_fotos_pdf(url):

        if not url:
            return []

        url = str(url).strip()

        if not url:
            return []

        try:

            # =================================================
            # FIELDSERVICE
            # =================================================

            if es_fieldservice(url):

                return descargar_fotos_fieldservice(
                    url
                )[:2]

            # =================================================
            # SIGOF
            # =================================================

            imagenes = extraer_imagenes(
                url
            )

            if not imagenes:
                return []

            fotos = []

            for imagen_url in imagenes[:2]:

                contenido = descargar_imagen_url(
                    imagen_url
                )

                if contenido:

                    fotos.append(
                        contenido
                    )

            return fotos[:2]

        except Exception:

            return []

    # ========================================================
    # ÁREA DISPONIBLE
    # ========================================================

    ancho_util = (
        ancho_pagina
        - doc.leftMargin
        - doc.rightMargin
    )

    alto_util = (
        alto_pagina
        - doc.topMargin
        - doc.bottomMargin
    )

    # ========================================================
    # RECORRER REGISTROS
    # ========================================================

    total_registros = len(df)

    for indice, (_, fila) in enumerate(
        df.iterrows()
    ):

        # ====================================================
        # DATOS
        # ====================================================

        suministro = obtener_valor(
            fila,
            col_suministro
        )

        medidor = obtener_valor(
            fila,
            col_medidor
        )

        direccion = obtener_valor(
            fila,
            col_direccion
        )

        obs = obtener_valor(
            fila,
            col_obs
        )

        obs_descripcion = obtener_valor(
            fila,
            col_obs_descripcion
        )

        lectura = obtener_valor(
            fila,
            col_lectura
        )

        # ====================================================
        # OBTENER URL REAL DEL HYPERLINK
        # ====================================================

        url_foto = obtener_valor(
            fila,
            "__url_foto"
        )

        # ====================================================
        # OBTENER FOTOS
        # ====================================================

        fotos = obtener_fotos_pdf(
            url_foto
        )

        # ====================================================
        # PREPARAR FOTOS
        # ====================================================

        fotos_pdf = []

        for contenido in fotos:

            try:

                imagen_original = PILImage.open(
                    BytesIO(contenido)
                )

                imagen_original.load()

                ancho_original = imagen_original.width

                alto_original = imagen_original.height

                # --------------------------------------------
                # CONVERTIR A RGB
                # --------------------------------------------

                if imagen_original.mode != "RGB":

                    if "A" in imagen_original.getbands():

                        fondo = PILImage.new(
                            "RGB",
                            imagen_original.size,
                            "white"
                        )

                        fondo.paste(
                            imagen_original,
                            mask=imagen_original.getchannel("A")
                        )

                        imagen_original = fondo

                    else:

                        imagen_original = (
                            imagen_original.convert(
                                "RGB"
                            )
                        )

                # --------------------------------------------
                # JPEG EN MEMORIA
                # --------------------------------------------

                buffer = BytesIO()

                imagen_original.save(
                    buffer,
                    format="JPEG",
                    quality=92
                )

                buffer.seek(0)

                fotos_pdf.append(
                    {
                        "buffer": buffer,
                        "ancho": ancho_original,
                        "alto": alto_original
                    }
                )

            except Exception:

                continue

        # ====================================================
        # MÁXIMO 2
        # ====================================================

        fotos_pdf = fotos_pdf[:2]

        cantidad_fotos = len(
            fotos_pdf
        )

        # ====================================================
        # ANCHOS BASE DE DATOS
        # ====================================================

        anchos_datos = [
            2.5 * cm,
            2.7 * cm,
            4.0 * cm,
            2.2 * cm,
            4.0 * cm,
            2.7 * cm
        ]

        ancho_datos = sum(
            anchos_datos
        )

        ancho_fotos = max(
            1 * cm,
            ancho_util - ancho_datos
        )

        # ====================================================
        # DISTRIBUIR ANCHO DE FOTOS
        # ====================================================

        if cantidad_fotos == 2:

            proporcion1 = (
                fotos_pdf[0]["ancho"]
                / fotos_pdf[0]["alto"]
            )

            proporcion2 = (
                fotos_pdf[1]["ancho"]
                / fotos_pdf[1]["alto"]
            )

            suma = (
                proporcion1
                + proporcion2
            )

            ancho_foto1 = (
                ancho_fotos
                * proporcion1
                / suma
            )

            ancho_foto2 = (
                ancho_fotos
                * proporcion2
                / suma
            )

        elif cantidad_fotos == 1:

            ancho_foto1 = ancho_fotos
            ancho_foto2 = 0

        else:

            ancho_foto1 = 0
            ancho_foto2 = 0

        # ====================================================
        # CALCULAR ALTURA
        # ====================================================

        alturas = []

        if cantidad_fotos >= 1:

            alturas.append(
                ancho_foto1
                * fotos_pdf[0]["alto"]
                / fotos_pdf[0]["ancho"]
            )

        if cantidad_fotos >= 2:

            alturas.append(
                ancho_foto2
                * fotos_pdf[1]["alto"]
                / fotos_pdf[1]["ancho"]
            )

        if alturas:

            alto_fotos = max(
                alturas
            )

        else:

            alto_fotos = 2.5 * cm

        # ====================================================
        # ALTURA MÁXIMA
        # ====================================================

        alto_encabezado = 1.1 * cm

        alto_maximo = (
            alto_util
            - alto_encabezado
            - 0.5 * cm
        )

        if alto_fotos > alto_maximo:

            factor = (
                alto_maximo
                / alto_fotos
            )

            ancho_foto1 *= factor
            ancho_foto2 *= factor

            alto_fotos = alto_maximo

        # ====================================================
        # CREAR IMÁGENES
        # ====================================================

        celdas_foto = []

        for posicion in range(2):

            if posicion >= cantidad_fotos:

                celdas_foto.append("")
                continue

            foto = fotos_pdf[posicion]

            ancho_original = foto["ancho"]
            alto_original = foto["alto"]

            if posicion == 0:

                ancho_celda = ancho_foto1

            else:

                ancho_celda = ancho_foto2

            ancho_maximo = max(
                1,
                ancho_celda - 4
            )

            alto_maximo_imagen = max(
                1,
                alto_fotos - 4
            )

            escala = min(
                ancho_maximo / ancho_original,
                alto_maximo_imagen / alto_original
            )

            ancho_final = (
                ancho_original
                * escala
            )

            alto_final = (
                alto_original
                * escala
            )

            imagen_pdf = RLImage(
                foto["buffer"],
                width=ancho_final,
                height=alto_final
            )

            celdas_foto.append(
                imagen_pdf
            )

        # ====================================================
        # ANCHOS FINALES
        # ====================================================

        anchos = [
            anchos_datos[0],
            anchos_datos[1],
            anchos_datos[2],
            anchos_datos[3],
            anchos_datos[4],
            anchos_datos[5],
            (
                ancho_foto1
                if cantidad_fotos >= 1
                else 3 * cm
            ),
            (
                ancho_foto2
                if cantidad_fotos >= 2
                else 3 * cm
            )
        ]

        # ====================================================
        # ASEGURAR QUE NO SUPERE A3
        # ====================================================

        suma_anchos = sum(
            anchos
        )

        if suma_anchos > ancho_util:

            factor = (
                ancho_util
                / suma_anchos
            )

            anchos = [
                x * factor
                for x in anchos
            ]

        # ====================================================
        # TABLA
        # ====================================================

        datos = [

            [
                Paragraph(
                    "SUMINISTRO",
                    estilo_encabezado
                ),
                Paragraph(
                    "MEDIDOR",
                    estilo_encabezado
                ),
                Paragraph(
                    "DIRECCIÓN",
                    estilo_encabezado
                ),
                Paragraph(
                    "OBS",
                    estilo_encabezado
                ),
                Paragraph(
                    "OBS_DESCRIPCION",
                    estilo_encabezado
                ),
                Paragraph(
                    "LECTURA",
                    estilo_encabezado
                ),
                Paragraph(
                    "FOTO 1",
                    estilo_encabezado
                ),
                Paragraph(
                    "FOTO 2",
                    estilo_encabezado
                )
            ],

            [
                Paragraph(
                    suministro,
                    estilo_dato
                ),
                Paragraph(
                    medidor,
                    estilo_dato
                ),
                Paragraph(
                    direccion,
                    estilo_dato
                ),
                Paragraph(
                    obs,
                    estilo_dato
                ),
                Paragraph(
                    obs_descripcion,
                    estilo_dato
                ),
                Paragraph(
                    lectura,
                    estilo_dato
                ),
                celdas_foto[0],
                celdas_foto[1]
            ]
        ]

        tabla = Table(
            datos,
            colWidths=anchos,
            rowHeights=[
                alto_encabezado,
                alto_fotos
            ],
            splitByRow=0,
            hAlign="CENTER"
        )

        # ====================================================
        # ESTILO
        # ====================================================

        tabla.setStyle(
            TableStyle([
                (
                    "BACKGROUND",
                    (0, 0),
                    (-1, 0),
                    colors.HexColor("#D9E1F2")
                ),
                (
                    "FONTNAME",
                    (0, 0),
                    (-1, 0),
                    "Helvetica-Bold"
                ),
                (
                    "FONTSIZE",
                    (0, 0),
                    (-1, 0),
                    8
                ),
                (
                    "ALIGN",
                    (0, 0),
                    (-1, -1),
                    "CENTER"
                ),
                (
                    "VALIGN",
                    (0, 0),
                    (-1, -1),
                    "MIDDLE"
                ),
                (
                    "GRID",
                    (0, 0),
                    (-1, -1),
                    0.7,
                    colors.black
                ),
                (
                    "LEFTPADDING",
                    (0, 0),
                    (-1, -1),
                    2
                ),
                (
                    "RIGHTPADDING",
                    (0, 0),
                    (-1, -1),
                    2
                ),
                (
                    "TOPPADDING",
                    (0, 0),
                    (-1, -1),
                    2
                ),
                (
                    "BOTTOMPADDING",
                    (0, 0),
                    (-1, -1),
                    2
                )
            ])
        )

        # ====================================================
        # AGREGAR
        # ====================================================

        elementos.append(
            tabla
        )

        # ====================================================
        # UNA PÁGINA POR REGISTRO
        # ====================================================

        if indice < total_registros - 1:

            elementos.append(
                PageBreak()
            )

    # ========================================================
    # GENERAR PDF
    # ========================================================

    doc.build(
        elementos
    )

    archivo_salida.seek(0)

    return archivo_salida