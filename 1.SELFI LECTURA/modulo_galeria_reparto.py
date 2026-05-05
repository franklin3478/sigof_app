import streamlit as st
import requests
import pandas as pd
from io import BytesIO
from openpyxl import load_workbook
import re
from datetime import datetime
from datetime import timedelta   

# ----------------------------
# UNIDADES
# ----------------------------
mapa_unidades = {
    "Valle Mantaro": {"idempresa": "4", "iduunn": "83"},
    "Tarma": {"idempresa": "4", "iduunn": "79"},
    "Huancavelica": {"idempresa": "4", "iduunn": "78"},
    "Ayacucho": {"idempresa": "4", "iduunn": "76"},
    "Selva Central": {"idempresa": "4", "iduunn": "80"},
    "Huánuco": {"idempresa": "4", "iduunn": "82"},
    "Pasco": {"idempresa": "4", "iduunn": "81"},
    "Huancayo": {"idempresa": "4", "iduunn": "77"},
    "Tingo María": {"idempresa": "4", "iduunn": "84"},
}

def ejecutar_galeria_reparto():

    hoy = datetime.today().strftime("%Y-%m-%d")

    # ----------------------------
    # SESSION STATE
    # ----------------------------
    if "login" not in st.session_state:
        st.session_state.login = False

    if "session" not in st.session_state:
        st.session_state.session = None

    if "ciclos" not in st.session_state:
        st.session_state.ciclos = {}

    if "df_galeria" not in st.session_state:
        st.session_state.df_galeria = None

    if "unidad_anterior" not in st.session_state:
        st.session_state.unidad_anterior = None

    if "total_ciclos" not in st.session_state:
        st.session_state.total_ciclos = None

    if "ciclos_previos" not in st.session_state:
        st.session_state.ciclos_previos = []

    if "reset_counter" not in st.session_state:
        st.session_state.reset_counter = 0

    if "observaciones" not in st.session_state:
        st.session_state.observaciones = {}

    if "data_signature" not in st.session_state:
        st.session_state.data_signature = None

    if "pagina" not in st.session_state:
        st.session_state.pagina = 1

    if "reset_page" not in st.session_state:
        st.session_state.reset_page = False

    # 🔥 NUEVO FIX: snapshot render (clave rendimiento + sincronía UI)
    if "render_token" not in st.session_state:
        st.session_state.render_token = 0

    # ----------------------------
    # LOGIN
    # ----------------------------
    if not st.session_state.login:

        st.title("🔐 SIGOF - Galería Reparto")

        user = st.text_input("Usuario SIGOF")
        password = st.text_input("Contraseña SIGOF", type="password")

        if st.button("Ingresar"):

            session = requests.Session()

            login_url = "http://sigof.distriluz.com.pe/plus/usuario/login"

            response = session.post(login_url, data={
                "data[Usuario][usuario]": user,
                "data[Usuario][pass]": password
            })

            if response.status_code == 200 and ("Salir" in response.text or "logout" in response.text.lower()):
                st.session_state.session = session
                st.session_state.login = True
                st.rerun()
            else:
                st.error("❌ Usuario o contraseña incorrectos")

    else:

        st.title("📷 Galería de Reparto SIGOF")

        session = st.session_state.session

        col1, col2 = st.columns(2)

        with col1:
            unidad = st.selectbox("Unidad de Negocio", list(mapa_unidades.keys()))
            u = mapa_unidades[unidad]

        if unidad != st.session_state.unidad_anterior:
            st.session_state.ciclos = {}
            st.session_state.df_galeria = None
            st.session_state.total_ciclos = None
            st.session_state.unidad_anterior = unidad

        with col2:

            ciclos_ids = []

            if st.session_state.ciclos:

                ciclos_seleccionados = st.multiselect(
                    "Ciclo",
                    options=list(st.session_state.ciclos.values()),
                    key=f"ciclo_select_{st.session_state.reset_counter}"
                )

                if set(ciclos_seleccionados) != set(st.session_state.ciclos_previos):
                    st.session_state.df_galeria = None
                    st.session_state.ciclos_previos = ciclos_seleccionados

                ciclos_ids = [
                    k for k, v in st.session_state.ciclos.items()
                    if v in ciclos_seleccionados
                ]

            else:
                st.selectbox("Ciclo", ["-- obtener ciclos primero --"], disabled=True)

        if st.session_state.get("total_ciclos"):
            st.info(f"🔎 {st.session_state.total_ciclos} ciclos encontrados")
        
        if st.button("📡 Obtener ciclos"):

            session.post(
                "http://sigof.distriluz.com.pe/plus/usuario/ajax_cambiar_sesion",
                data={"idempresa": u["idempresa"], "iduunn": u["iduunn"]}
            )

            session.get("http://sigof.distriluz.com.pe/plus/ComrepOrdenrepartos/ordenes_historico")

            url = f"http://sigof.distriluz.com.pe/plus/ComrepOrdenrepartos/ajax_listar_tabla_repartos_historico/U/0/0/0/0/{hoy}/{hoy}/0/"

            r = session.get(url, headers={"X-Requested-With": "XMLHttpRequest"})

            matches = re.findall(r'(\d+)-(Ciclo[^<"]+)', r.text)

            ciclos = {}

            for idc, desc in matches:
                ciclos[idc] = f"{idc}-{desc}".strip()

            st.session_state.ciclos = ciclos
            st.session_state.df_galeria = None
            st.session_state.total_ciclos = len(ciclos)

            st.session_state.ciclos_previos = []
            st.session_state.reset_counter += 1
            
            st.session_state.render_token += 1   # 🔥 FIX

            st.rerun()

        def descargar_datos(session, idciclo):

            url = f"http://sigof.distriluz.com.pe/plus/ComrepOrdenrepartos/ajax_reporte_excel_ordenes_historico/U/0/{idciclo}/0/0/{hoy}/{hoy}/0/"

            r = session.get(url)

            if r.status_code == 200 and r.content:
                wb = load_workbook(BytesIO(r.content))
                ws = wb.active

                data = []
                for row in ws.iter_rows(values_only=True):
                    data.append(row)

                return pd.DataFrame(data[1:], columns=data[0])

            return None

        if ciclos_ids and st.button("📥 Descargar datos"):

            dfs = []

            for cid in ciclos_ids:
                df_temp = descargar_datos(session, cid)
                if df_temp is not None:
                    dfs.append(df_temp)

            if dfs:
                df = pd.concat(dfs, ignore_index=True)

                new_signature = f"{len(df)}_{sorted(ciclos_ids)}"

                if st.session_state.data_signature != new_signature:
                    st.session_state.observaciones = {}

                st.session_state.data_signature = new_signature
                st.session_state.df_galeria = df
                
                st.session_state.render_token += 1  # 🔥 FIX

        df = st.session_state.df_galeria
        
        if df is not None and "suministro" in df.columns:

            df_base = df.copy()

            # ✅ CREAR HORA UNA SOLA VEZ
            if "fecha_ejecutado" in df_base.columns:
                df_base["fecha_ejecutado"] = pd.to_datetime(df_base["fecha_ejecutado"], errors="coerce")                
            else:
                df_base["hora"] = None
                        
            colf1, colf2, colf3, colf4 = st.columns(4)
            
            ciclo_visual = st.session_state.get("ciclo_visual", "Todos")
            if ciclo_visual != "Todos":
                df_base = df_base[df_base["idciclo"] == ciclo_visual]

            filtro_lecturista = st.session_state.get("filtro_lecturista", "Todos")
            filtro_sector = st.session_state.get("filtro_sector", "Todos")
            filtro_ruta = st.session_state.get("filtro_ruta", "Todos")

            df_temp = df_base.copy()
            if filtro_sector != "Todos":
                df_temp = df_temp[df_temp["sector"] == filtro_sector]
            if filtro_ruta != "Todos":
                df_temp = df_temp[df_temp["ruta"] == filtro_ruta]

            with colf1:
                lecturistas = ["Todos"] + sorted(df_temp["lecturista"].dropna().unique().tolist())
                filtro_lecturista = st.selectbox("Lecturista", lecturistas, key="filtro_lecturista")

            df_temp = df_base.copy()
            if filtro_lecturista != "Todos":
                df_temp = df_temp[df_temp["lecturista"] == filtro_lecturista]
            if filtro_ruta != "Todos":
                df_temp = df_temp[df_temp["ruta"] == filtro_ruta]

            with colf2:
                sectores = ["Todos"] + sorted(df_temp["sector"].dropna().unique().tolist())
                filtro_sector = st.selectbox("Sector", sectores, key="filtro_sector")

            df_temp = df_base.copy()
            if filtro_lecturista != "Todos":
                df_temp = df_temp[df_temp["lecturista"] == filtro_lecturista]
            if filtro_sector != "Todos":
                df_temp = df_temp[df_temp["sector"] == filtro_sector]

            with colf3:
                rutas = ["Todos"] + sorted(df_temp["ruta"].dropna().unique().tolist())
                filtro_ruta = st.selectbox("Ruta", rutas, key="filtro_ruta")

            with colf4:           
                fotos_por_pagina = st.selectbox("Fotos por página", [50, 100, 200], index=1)
            
            # 🔥 APLICAR FILTROS ANTES DEL SLIDER
            df_filtros = df_base.copy()

            if ciclo_visual != "Todos":
                df_filtros = df_filtros[df_filtros["idciclo"] == ciclo_visual]

            if filtro_lecturista != "Todos":
                df_filtros = df_filtros[df_filtros["lecturista"] == filtro_lecturista]

            if filtro_sector != "Todos":
                df_filtros = df_filtros[df_filtros["sector"] == filtro_sector]

            if filtro_ruta != "Todos":
                df_filtros = df_filtros[df_filtros["ruta"] == filtro_ruta]

            # 🔥 LIMPIAR FECHA SOLO AQUÍ
            df_filtros["fecha_ejecutado"] = pd.to_datetime(df_filtros["fecha_ejecutado"], errors="coerce")
            df_filtros = df_filtros.dropna(subset=["fecha_ejecutado"])

            if df_filtros.empty:
                st.warning("No hay datos con los filtros seleccionados")
                return

            df_filtros = df_filtros.sort_values("fecha_ejecutado")

            # 🔥 RANGO REAL YA FILTRADO
            fecha_min = df_filtros["fecha_ejecutado"].iloc[0]
            fecha_max = df_filtros["fecha_ejecutado"].iloc[-1]

            # 🔥 SLIDER CORRECTO
            rango_fechas = st.slider(
                "⏱ Línea de tiempo",
                min_value=fecha_min.to_pydatetime(),
                max_value=fecha_max.to_pydatetime(),
                value=(fecha_min.to_pydatetime(), fecha_max.to_pydatetime()),
                format="DD/MM HH:mm",
                step=timedelta(minutes=1)
            )

            df_filtrado = df_filtros.copy()

            if filtro_lecturista != "Todos":
                df_filtrado = df_filtrado[df_filtrado["lecturista"] == filtro_lecturista]

            if filtro_sector != "Todos":
                df_filtrado = df_filtrado[df_filtrado["sector"] == filtro_sector]

            if filtro_ruta != "Todos":
                df_filtrado = df_filtrado[df_filtrado["ruta"] == filtro_ruta]

            fecha_inicio, fecha_fin = rango_fechas

            df_filtrado = df_filtrado[
                (df_filtrado["fecha_ejecutado"] >= fecha_inicio) &
                (df_filtrado["fecha_ejecutado"] <= fecha_fin)
            ]
            
            uunn = u["iduunn"]

            df_filtrado["url"] = df_filtrado.apply(
                lambda row: f"https://d3jgwc2y5nosue.cloudfront.net/repartos/{str(row['pfactura']).strip()}/{uunn}/{str(row['pfactura']).strip()}_{uunn}_{str(row['suministro']).strip()}_RECIBO.png"
                if str(row.get("foto")).strip().lower() == "ver foto"
                else None,
                axis=1
            )

            urls = [
                (row["url"], str(row["suministro"]).strip(), row.get("lecturista", ""), row.get("idciclo"), row.get("observacion", ""))
                for _, row in df_filtrado.dropna(subset=["url"]).iterrows()
            ]

            if urls:

                ciclos_con_foto = sorted(list(set([x[3] for x in urls if x[3]])))
                opciones = ["Todos"] + ciclos_con_foto

                ciclo_visual = st.radio(
                    "Filtrar por ciclo",
                    options=opciones,
                    horizontal=True,
                    key="ciclo_visual"
                )

                if ciclo_visual != "Todos":
                    urls = [x for x in urls if x[3] == ciclo_visual]

                # 🔥 FIX: snapshot por render evita desincronización visual
                render_id = f"{st.session_state.render_token}_{st.session_state.pagina}_{len(urls)}"

                colp1, colp2 = st.columns([2, 3])

                max_page = max(1, (len(urls) - 1) // fotos_por_pagina + 1)

                # 🔥 CLAMP del valor antes del widget (ESTO ES LO CLAVE)
                if "pagina" in st.session_state:
                    if st.session_state.pagina > max_page:
                        st.session_state.pagina = 1

                with colp1:
                    pagina = st.number_input(
                        "📄 Página",
                        1,
                        max(1, (len(urls) // fotos_por_pagina) + 1),
                        value=st.session_state.get("pagina", 1),
                        key="pagina"    
                    )
                with colp2:
                    busqueda = st.text_input("🔎 Buscar suministro", key=f"busqueda_{render_id}")

                if busqueda:
                    urls = [x for x in urls if busqueda.lower() in x[1].lower()]
                    pagina = 1
                                
                inicio = (pagina - 1) * fotos_por_pagina
                fin = inicio + fotos_por_pagina

                urls_pagina = urls[inicio:fin]

                st.success(f"📸 Fotos encontradas: {len(urls)}")
                st.subheader("📷 Galería de Fotos")

                opciones_obs = [
                    "CORRECTO",
                    "SOLO RECIBO/PUERTA",
                    "FOTOGRAFIA INCORRECTO/BORROSO/DESENFOCADO",
                    "SOLO RECIBO",
                    "SOLO SUMINISTRO",
                    "DATOS NO COINCIDEN",
                    "NI RECIBO/NI SUMINISTRO"
                ]

                for i, (url, suministro, lecturista, _, observacion) in enumerate(urls_pagina):

                    if i % 5 == 0:
                        if i != 0:
                            st.markdown("<hr>", unsafe_allow_html=True)
                        cols = st.columns(5)

                    with cols[i % 5]:

                        st.image(url, use_container_width=True)

                        obs = "" if pd.isna(observacion) else str(observacion)
                                                
                        st.markdown(f"<div style='text-align:center;font-size:13px'>"f"{suministro}<br>"f"{obs if obs else ''}<br>"f"{lecturista}"f"</div>",unsafe_allow_html=True)

                        st.markdown("<div style='text-align:center;font-weight:bold;'>Observaciones</div>", unsafe_allow_html=True)

                        clave = f"{pagina}_{suministro}"
                        valor = st.session_state.observaciones.get(clave, "CORRECTO")

                        opciones = [""] + opciones_obs

                        seleccion = st.selectbox(
                            "",
                            opciones,
                            index=opciones.index(valor) if valor in opciones else opciones.index("CORRECTO"),
                            key=f"obs_{clave}"
                        )

                        st.session_state.observaciones[clave] = seleccion

        if st.session_state.df_galeria is not None and not st.session_state.df_galeria.empty:

            # 🔥 Copia directa del dataframe ya cargado
            df_export = st.session_state.df_galeria.copy()

            # ⚡ Agregar observaciones desde session_state (MUY RÁPIDO)
            df_export["obs_extra"] = df_export["suministro"].astype(str).map(
                st.session_state.observaciones
            ).fillna("")

            # 🧠 Generar archivo en memoria (sin tocar SIGOF)
            output = BytesIO()

            with pd.ExcelWriter(output, engine="openpyxl") as writer:
                df_export.to_excel(writer, index=False)

            # 📥 Descarga directa (1 solo botón, sin pasos extra)
            st.download_button(
                "📥 Descargar Excel con observaciones",
                data=output.getvalue(),
                file_name="SIGOF_reporte.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
            
        if st.button("Cerrar sesión"):
            st.session_state.clear()
            st.rerun()