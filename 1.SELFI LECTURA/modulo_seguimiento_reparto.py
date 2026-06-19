import streamlit as st
import requests
from datetime import datetime
from bs4 import BeautifulSoup
import pandas as pd
import re
from io import BytesIO
from concurrent.futures import ThreadPoolExecutor, as_completed
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import math
import numpy as np
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

def limpiar_datos_unidad():
    keys_a_limpiar = [
        "ciclos",
        "monitoreo_df",
        "resumen_reparto",
    ]

    for k in keys_a_limpiar:
        st.session_state.pop(k, None)

# ----------------------------
# UNIDADES
# ----------------------------
mapa_unidades = {
    "Valle Mantaro": {"idempresa": "4", "iduunn": "83"},
    "Tarma": {"idempresa": "4", "iduunn": "79"},
    "Huancavelica": {"idempresa": "4", "iduunn": "78"},
    "Ayacucho": {"idempresa": "4", "iduunn": "76"},
    "Selva Central": {"idempresa": "4", "iduunn": "80"},
    "Huánuco": {"idempresa": "4", "iduunn": "82"},
    "Pasco": {"idempresa": "4", "iduunn": "81"},
    "Huancayo": {"idempresa": "4", "iduunn": "77"},
    "Tingo María": {"idempresa": "4", "iduunn": "84"},
}

# ----------------------------
# HELPERS
# ----------------------------
def cambiar_unidad(session, unidad):
    try:
        session.post(
            "http://sigof.distriluz.com.pe/plus/usuario/ajax_cambiar_sesion",
            data=mapa_unidades[unidad],
            timeout=20
        )
    except requests.exceptions.RequestException as e:
        st.error(f"Error cambiando unidad: {e}")

# ----------------------------
# LOGIN
# ----------------------------
def login_sigof():
    st.subheader("🔐 Login SIGOF")

    usuario = st.text_input("Usuario")
    password = st.text_input("Contraseña", type="password")

    if st.button("Ingresar"):
        session = requests.Session()
        try:
            r = session.post(
                "http://sigof.distriluz.com.pe/plus/usuario/login",
                data={
                    "data[Usuario][usuario]": usuario,
                    "data[Usuario][pass]": password
                },
                timeout=30
            )

            if "Salir" in r.text:
                st.session_state.session_sigof = session
                st.session_state.logueado_sigof = True
                st.rerun()
            else:
                st.error("Credenciales incorrectas")

        except requests.exceptions.RequestException as e:
            st.error(f"Error de conexión: {e}")

# ----------------------------
# CICLOS
# ----------------------------
def obtener_ciclos(session, unidad, fecha):
    cambiar_unidad(session, unidad)

    url = f"http://sigof.distriluz.com.pe/plus/ComrepOrdenrepartos/ajax_listar_tabla_repartos_historico/U/0/0/0/0/{fecha}/{fecha}/0/"

    try:
        r = session.get(url, headers={"X-Requested-With": "XMLHttpRequest"}, timeout=30)
        matches = re.findall(r'(\d+)-(Ciclo[^"<]+)', r.text)
        return {idc: f"{idc} - {desc.strip()}" for idc, desc in matches}
    except requests.exceptions.RequestException as e:
        st.error(f"Error obteniendo ciclos: {e}")
        return {}

# ----------------------------
# MONITOREO
# ----------------------------
def obtener_monitoreo(session, unidad, fecha, idciclo):
    cambiar_unidad(session, unidad)

    url = f"http://sigof.distriluz.com.pe/plus/ComrepOrdenrepartos/ajax_monitorear_reparto_reload/U/{fecha}/{fecha}/{idciclo}/0/0"

    try:
        r = session.get(url, headers={"X-Requested-With": "XMLHttpRequest"}, timeout=30)
        return r.text
    except requests.exceptions.RequestException as e:
        st.error(f"Error en monitoreo: {e}")
        return ""

def parsear_monitoreo(html):
    soup = BeautifulSoup(html, "html.parser")
    tabla = soup.find("table", {"id": "list-monitorear-lecturistas"})

    if tabla is None:
        return pd.DataFrame()

    filas = tabla.find("tbody").find_all("tr")
    data = []

    for fila in filas:
        cols = fila.find_all("td")
        if len(cols) < 18:
            continue
        
        avance_raw = cols[14].get_text(strip=True)
        avance_raw = re.sub(r"\s+", "", avance_raw)

        match_avance = re.search(r"(\d+)%", avance_raw)
        if match_avance:
            avance_raw = f"{match_avance.group(1)}%"

        data.append({
            "repartidor": cols[8].get_text(strip=True),
            "asignado": cols[9].get_text(strip=True),
            "descarga": cols[10].get_text(strip=True),
            "% avance": avance_raw,
            "fin": cols[12].get_text(strip=True),
            "pendiente": cols[13].get_text(strip=True),
            "entregado": cols[16].get_text(strip=True),
            "paso_ruta": cols[17].get_text(strip=True),
        })

    return pd.DataFrame(data)

# ----------------------------
# DESCARGA EXCEL
# ----------------------------
def descargar_excel_ciclo(session, unidad, idc):
    cambiar_unidad(session, unidad)

    hoy = datetime.today().strftime("%Y-%m-%d")

    url = f"http://sigof.distriluz.com.pe/plus/ComrepOrdenrepartos/ajax_reporte_excel_ordenes_historico/U/0/{idc}/0/0/{hoy}/{hoy}/0/"

    try:
        r = session.get(url, timeout=60)

        if r.status_code == 200 and r.content:
            return load_workbook(BytesIO(r.content))
    except requests.exceptions.RequestException:
        pass

    return None

def descargar_ciclos_excel(session, unidad, ids):

    resultados = {}

    with ThreadPoolExecutor(max_workers=4) as executor:
        futures = {
            executor.submit(descargar_excel_ciclo, session, unidad, idc): idc
            for idc in ids
        }

        for future in as_completed(futures):
            idc = futures[future]
            wb = future.result()

            if wb:
                resultados[idc] = wb

    return resultados

# ----------------------------
# EXCEL → DF
# ----------------------------
def excel_a_dataframe(wb):
    ws = wb.active

    foto_col_idx = None
    for idx, cell in enumerate(ws[1], start=1):
        if str(cell.value).strip().lower() == "foto":
            foto_col_idx = idx
            break

    foto_links = {}

    if foto_col_idx:
        foto_col_letter = get_column_letter(foto_col_idx)

        for cell in ws[foto_col_letter]:
            if cell.hyperlink:
                foto_links[cell.row] = cell.hyperlink.target

    data = list(ws.values)
    df = pd.DataFrame(data)

    df.columns = df.iloc[0]
    df = df[1:]
    df.columns = df.columns.str.strip()

    if "foto" in df.columns:
        df["foto"] = df.apply(
            lambda row: foto_links.get(row.name + 1, row["foto"]),
            axis=1
        )
        df["foto"] = df["foto"].astype(str)

    return df

# ----------------------------
# LIMPIEZA
# ----------------------------
def limpiar_dataframe(df):
    df.columns = df.columns.str.strip().str.lower()

    df["fise"] = pd.to_numeric(df.get("fise", 0), errors="coerce").fillna(0)

    df["foto"] = (
        df.get("foto", "")
        .astype(str)
        .str.strip()
        .replace(["nan", "none", "NaN", ""], "")
    )

    return df

# 🔥 NUEVA FUNCIÓN AQUÍ
def obtener_df_total(session, unidad, ids_ciclos, ciclos_dict):

    cache_key = (unidad, tuple(sorted(ids_ciclos)))

    if st.session_state.get("df_total_key") == cache_key:
        return st.session_state.get("df_total")

    # 🔥 recalcular porque cambió selección
    wbs = descargar_ciclos_excel(session, unidad, ids_ciclos)

    dfs = []
    for idc in ids_ciclos:
        wb = wbs.get(idc)
        if wb:
            df_tmp = excel_a_dataframe(wb)
            df_tmp["ciclo"] = ciclos_dict[idc]
            dfs.append(df_tmp)

    if not dfs:
        return None

    df_total = pd.concat(dfs, ignore_index=True)
    df_total = limpiar_dataframe(df_total)

    # 🔥 guardar con clave
    st.session_state.df_total = df_total
    st.session_state.df_total_key = cache_key

    return df_total

# ----------------------------
# FLAGS + FEATURES
# ----------------------------
def agregar_flags(df):
    df["fise_es_1"] = df["fise"] == 1
    df["tiene_foto"] = df["foto"].str.contains("http", na=False)
    df["fise_con_foto_flag"] = df["fise_es_1"] & df["tiene_foto"]
    return df

# ----------------------------
# PROCESAMIENTO
# ----------------------------
def procesar_dataframe(df):
    df = agregar_flags(df)

    def calc(total):
        return math.ceil(total * 0.10) if total < 10 else math.floor(total * 0.10)

    resumen = (
        df.groupby(["ruta", "lecturista", "ciclo"])
        .agg(
            total_suministros=("ruta", "count"),
            fise_obligatorio=("fise_es_1", "sum"),
            fotos_tomadas=("tiene_foto", "sum"),
            fise_con_foto=("fise_con_foto_flag", "sum")
        )
        .reset_index()
    )

    st.session_state.lecturistas_options = sorted(resumen["lecturista"].dropna().unique())
    st.session_state.ciclos_options = sorted(resumen["ciclo"].dropna().unique())
    st.session_state.rutas_options = sorted(resumen["ruta"].dropna().unique())

    resumen["fise_sin_foto"] = (resumen["fise_obligatorio"] - resumen["fise_con_foto"]).clip(lower=0)
        
    resumen["cant_min_foto"] = np.where(
        resumen["total_suministros"] < 10,
        np.ceil(resumen["total_suministros"] * 0.10),
        np.floor(resumen["total_suministros"] * 0.10)
    )

    resumen["consignado"] = "NO CUMPLIÓ EL 10%"

    resumen.loc[resumen["fotos_tomadas"] == 0, "consignado"] = "NO CUENTA CON FOTOS"

    resumen.loc[
        resumen["fotos_tomadas"] >= resumen["cant_min_foto"],
        "consignado"
    ] = "CUMPLIÓ EL 10%"

    resumen["cumplimiento_fise"] = resumen.apply(
        lambda r: "CUMPLIÓ" if r["fise_con_foto"] == r["fise_obligatorio"] else "NO CUMPLIÓ",
        axis=1
    )

    return resumen

def to_excel(df):
    output = BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df.to_excel(writer, index=False)
    return output.getvalue()

def to_excel_unificado(df_monitoreo, df_fise, df_10):
    output = BytesIO()

    with pd.ExcelWriter(output, engine="openpyxl") as writer:

        hojas = {
            "Monitoreo": df_monitoreo,
            "FISE": df_fise,
            "Cumplimiento": df_10
        }

        for nombre_hoja, df in hojas.items():

            df.to_excel(
                writer,
                sheet_name=nombre_hoja,
                index=False
            )

            ws = writer.sheets[nombre_hoja]

            # Encabezado en negrita
            for cell in ws[1]:

                # Convertir encabezado a MAYÚSCULAS
                if cell.value:
                    cell.value = str(cell.value).upper()

                # Letra blanca y negrita
                cell.font = Font(
                    bold=True,
                    color="FFFFFF"
                )

                # Fondo azul corporativo
                cell.fill = PatternFill(
                    fill_type="solid",
                    start_color="4472C4",
                    end_color="4472C4"
                )
            for column in ws.columns:

                max_length = 0
                column_letter = get_column_letter(column[0].column)

                for cell in column:
                    try:
                        if cell.value is not None:
                            max_length = max(max_length, len(str(cell.value)))
                    except:
                        pass

                ws.column_dimensions[column_letter].width = max_length + 4

    return output.getvalue()

def aplicar_filtros_global(df):
    lect = st.session_state.get("filtro_lecturista", [])
    ciclo = st.session_state.get("filtro_ciclo", [])
    
    if lect and "lecturista" in df.columns:
        df = df[df["lecturista"].isin(lect)]

    if ciclo and "ciclo" in df.columns:
        df = df[df["ciclo"].isin(ciclo)]
    
    return df

def resetear_filtros():
    keys = [
        "filtro_lecturista",
        "filtro_ciclo",
        "filtro_ruta",
        "filtro_ruta_local",
        "filtro_fise_local",
        "filtro_consignado_local",
    ]

    for k in keys:
        if k in st.session_state:
            del st.session_state[k]

    # 🔥 evita que streamlit recupere valores anteriores
    st.session_state["filtro_lecturista"] = []
    st.session_state["filtro_ciclo"] = []

# ----------------------------
# MAIN
# ----------------------------
unidad_actual = st.session_state.get("unidad_actual")
def ejecutar_seguimiento_reparto():

    st.session_state.setdefault("logueado_sigof", False)
    st.session_state.setdefault("ciclos", {})

    if not st.session_state.logueado_sigof:
        login_sigof()
        return

    session = st.session_state.session_sigof
    fecha = datetime.today().strftime("%Y-%m-%d")

    st.title("📊Seguimiento de Reparto")

    col1, col2 = st.columns(2)

    with col1:
        st.session_state.setdefault("unidad_actual", None)

        unidad = st.selectbox("Unidad", list(mapa_unidades.keys()))

        # 🔥 DETECTAR CAMBIO DE UNIDAD
        if st.session_state.get("unidad_actual") != unidad:
            limpiar_datos_unidad()
            st.session_state["unidad_actual"] = unidad
            st.rerun()
    
    # Botón NO se mueve (misma lógica)
    if st.button("📡 Obtener ciclos"):
        st.session_state.ciclos = obtener_ciclos(session, unidad, fecha)

        # 🔥 FORZAR RESET REAL
        st.session_state["ciclos_multiselect"] = []

        # limpiar data previa
        st.session_state.pop("monitoreo_df", None)
        st.session_state.pop("resumen_reparto", None)

        # limpiar filtros
        for k in [
            "filtro_lecturista",
            "filtro_ciclo",
            "filtro_ruta",
            "filtro_ruta_local",
            "filtro_fise_local",
            "filtro_consignado_local",
        ]:
            st.session_state.pop(k, None)

        st.rerun()

    ciclos_dict = st.session_state.get("ciclos", {})

    with col2:
        if ciclos_dict:
            ciclos_seleccionados = st.multiselect(
                "Ciclo",
                list(ciclos_dict.values()),
                default=st.session_state.get("ciclos_multiselect", []),
                key="ciclos_multiselect"
            )

            ids_ciclos = [
                list(ciclos_dict.keys())[list(ciclos_dict.values()).index(c)]
                for c in ciclos_seleccionados
            ]
        else:
            ciclos_seleccionados = []
            ids_ciclos = []

    if not ciclos_dict:
        st.info("Selecciona una unidad y obtén ciclos")
        return
    
    # ----------------------------
    # FILTROS GLOBALES (ARRIBA)
    # ----------------------------
    st.markdown("---")
    
    lecturistas_options = st.session_state.get("lecturistas_options", [])
    ciclos_options = st.session_state.get("ciclos_options", [])
    
    col1, col2 = st.columns(2)

   
    # ----------------------------
    # FILTROS DEPENDIENTES (CORRECTO LECTURISTA → CICLO)
    # ----------------------------

    df_base_global = st.session_state.get("resumen_reparto")

    if df_base_global is None:
        lecturistas_filtrados = st.session_state.get("lecturistas_options", [])
        ciclos_filtrados = st.session_state.get("ciclos_options", [])

    else:
        df_tmp = df_base_global.copy()

        lect_sel = st.session_state.get("filtro_lecturista", [])
        ciclo_sel = st.session_state.get("filtro_ciclo", [])

        # =========================
        # 🔥 LECTURISTA FILTRA BASE PARA CICLO
        # =========================
        if lect_sel:
            df_tmp_ciclo = df_tmp[df_tmp["lecturista"].isin(lect_sel)]
        else:
            df_tmp_ciclo = df_tmp

        # ciclos DEPENDEN del lecturista
        ciclos_filtrados = sorted(df_tmp_ciclo["ciclo"].dropna().unique())

        # =========================
        # 🔥 CICLO FILTRA BASE PARA LECTURISTA (opcional simétrico)
        # =========================
        if ciclo_sel:
            df_tmp_lect = df_tmp[df_tmp["ciclo"].isin(ciclo_sel)]
        else:
            df_tmp_lect = df_tmp

        lecturistas_filtrados = sorted(df_tmp_lect["lecturista"].dropna().unique())
        
    with col1:
        st.multiselect("Lecturista", lecturistas_filtrados, key="filtro_lecturista")

    with col2:
        st.multiselect("Ciclo", ciclos_filtrados, key="filtro_ciclo")

    if st.button("📊 Monitoreo"):

        if not ids_ciclos:
            st.warning("Selecciona al menos un ciclo")
            return

        df_monitoreo_total = []

        for idc in ids_ciclos:
            html = obtener_monitoreo(session, unidad, fecha, idc)
            df_mon = parsear_monitoreo(html)

            if not df_mon.empty:
                df_mon["ciclo"] = ciclos_dict[idc]
                df_monitoreo_total.append(df_mon)

        if df_monitoreo_total:
            df_monitoreo_total = pd.concat(df_monitoreo_total, ignore_index=True)            
            df_monitoreo_total.rename(columns={"repartidor": "lecturista"}, inplace=True)
            
            # 🔥 SIEMPRE recalcular fotos (SOLUCIÓN)
            df_total = obtener_df_total(session, unidad, ids_ciclos, ciclos_dict)

            if df_total is not None:

                resumen = procesar_dataframe(df_total)

                df_fotos_lecturista = (
                    resumen.groupby(["lecturista", "ciclo"], as_index=False)
                    .agg(total_fotos=("fotos_tomadas", "sum"))
                )

                st.session_state.resumen_fotos_lecturista = df_fotos_lecturista
           

            # -------------------------
            # MAPEAR FOTOS
            # -------------------------
            df_fotos_lecturista = st.session_state.get("resumen_fotos_lecturista")

            if df_fotos_lecturista is not None:

                # 🔥 crear mapa directo (más estable)
                foto_map = (
                    df_fotos_lecturista
                    .set_index(["lecturista", "ciclo"])["total_fotos"]
                )

                # 🔥 mapear con merge directo (evita errores de index)
                df_monitoreo_total = df_monitoreo_total.merge(
                    df_fotos_lecturista[["lecturista", "ciclo", "total_fotos"]],
                    on=["lecturista", "ciclo"],
                    how="left"
                )

                # renombrar directamente (evita columna extra problemática)
                df_monitoreo_total.rename(columns={"total_fotos": "foto"}, inplace=True)

                # limpiar valores
                df_monitoreo_total["foto"] = pd.to_numeric(
                    df_monitoreo_total["foto"],
                    errors="coerce"
                ).fillna(0)

            else:
                df_monitoreo_total["foto"] = 0
                        
            # =========================
            # 🔥 CONSIGNADO 10% GENERAL (SIN columna extra de cálculo)
            # =========================

            df_monitoreo_total["foto"] = pd.to_numeric(df_monitoreo_total["foto"], errors="coerce").fillna(0)
            df_monitoreo_total["asignado"] = pd.to_numeric(df_monitoreo_total["asignado"], errors="coerce").fillna(0)
            df_monitoreo_total["descarga"] = pd.to_numeric(df_monitoreo_total["descarga"], errors="coerce").fillna(0)
            df_monitoreo_total["% avance"] = df_monitoreo_total["% avance"].astype(str)

            df_monitoreo_total["% avance"] = np.where(
                df_monitoreo_total["% avance"].isin(["0%", "0 %", "0"]),
                "NO INICIÓ",
                df_monitoreo_total["% avance"]
            )
           
            df_monitoreo_total["descarga"] = np.where(
                df_monitoreo_total["descarga"] == 0,
                "NO DESCARGÓ",
                df_monitoreo_total["descarga"]
            )
            
            df_monitoreo_total["consignado"] = np.where(
                df_monitoreo_total["foto"] >= np.where(
                    df_monitoreo_total["asignado"] < 10,
                    1,
                    np.floor(df_monitoreo_total["asignado"] * 0.10)
                ),
                "CUMPLIÓ EL 10% GENERAL",
                "NO CUMPLIÓ EL 10% GENERAL"
            )

            cols = df_monitoreo_total.columns.tolist()

            if "ciclo" in cols and "lecturista" in cols:
                cols.remove("ciclo")
                idx = cols.index("lecturista")
                cols.insert(idx, "ciclo")
            
            if "foto" in cols and "consignado" in cols:
                cols.remove("foto")
                idx = cols.index("consignado")
                cols.insert(idx, "foto")

            df_monitoreo_total = df_monitoreo_total[cols]
            
            st.session_state.lecturistas_options = sorted(df_monitoreo_total["lecturista"].dropna().unique())
            st.session_state.ciclos_options = sorted(df_monitoreo_total["ciclo"].dropna().unique())
            st.session_state.rutas_options = []  # aún no existen en monitoreo
            
            resetear_filtros()  # 👈 LIMPIA FILTROS

            st.session_state.monitoreo_df = df_monitoreo_total
            
            if "resumen_reparto" in st.session_state:
                del st.session_state.resumen_reparto

            st.rerun()

        else:
            st.warning("No hay datos de monitoreo")

    if "monitoreo_df" in st.session_state:
        st.subheader("📡 Monitoreo")

        df_mon = st.session_state.monitoreo_df.copy()
        df_mon = aplicar_filtros_global(df_mon)
                
        st.dataframe(df_mon, use_container_width=True)
                   
        # ----------------------------
        # PROCESAR EXCEL SOLO AQUÍ
        # ----------------------------
        if "resumen_reparto" not in st.session_state:


            wbs = descargar_ciclos_excel(session, unidad, ids_ciclos)

            def procesar_wb(idc, wb):
                if wb:
                    df_tmp = excel_a_dataframe(wb)
                    df_tmp["idciclo"] = idc
                    df_tmp["ciclo"] = ciclos_dict[idc]
                    return df_tmp
                return None

            dfs = []

            with ThreadPoolExecutor(max_workers=4) as executor:
                futures = [
                    executor.submit(procesar_wb, idc, wbs.get(idc))
                    for idc in ids_ciclos
                ]

                for future in as_completed(futures):
                    df_tmp = future.result()
                    if df_tmp is not None:
                        dfs.append(df_tmp)

            if not dfs:
                st.error("No se pudo obtener datos de reportes")
                return

            df_total = obtener_df_total(session, unidad, ids_ciclos, ciclos_dict)

            if df_total is None:
                st.error("No se pudo obtener datos")
                return

            resumen = procesar_dataframe(df_total)
            st.session_state.resumen_reparto = resumen

            # =========================
            # TOTAL FOTOS POR LECTURISTA
            # =========================
            df_fotos_lecturista = (
                resumen.groupby("lecturista", as_index=False)
                .agg(
                    total_fotos=("fotos_tomadas", "sum"),
                    total_suministros=("total_suministros", "sum"),
                    fise_obligatorio=("fise_obligatorio", "sum"),   
                    fise_con_foto=("fise_con_foto", "sum"),
                )
            )

            df_fotos_lecturista["fise_sin_foto"] = (
                df_fotos_lecturista["fise_obligatorio"] - df_fotos_lecturista["fise_con_foto"]
            )

            st.session_state.resumen_fotos_lecturista = df_fotos_lecturista

    # ----------------------------
    # FILTROS
    # ----------------------------
    if "resumen_reparto" in st.session_state:

        st.markdown("---")

        # ----------------------------
        # BASE
        # ----------------------------
        df = st.session_state.resumen_reparto.copy()

        # aplicar SOLO filtros globales primero
        df = aplicar_filtros_global(df)
              
        # =========================
        # BASE PARA OPCIONES (NO incluye ruta)
        # =========================
        df_base_opciones = st.session_state.resumen_reparto.copy()

        lect = st.session_state.get("filtro_lecturista", [])
        ciclo = st.session_state.get("filtro_ciclo", [])

        if lect:
            df_base_opciones = df_base_opciones[df_base_opciones["lecturista"].isin(lect)]

        if ciclo:
            df_base_opciones = df_base_opciones[df_base_opciones["ciclo"].isin(ciclo)]

        rutas_filtradas = sorted(df_base_opciones["ruta"].dropna().unique())

        # =========================
        # BASE REAL (para filtros finales)
        # =========================
        df_base = st.session_state.resumen_reparto.copy()

        ruta = st.session_state.get("filtro_ruta_local", [])
        fise_sel_actual = st.session_state.get("filtro_fise_local", [])
        consignado_sel_actual = st.session_state.get("filtro_consignado_local", [])

        if lect:
            df_base = df_base[df_base["lecturista"].isin(lect)]

        if ciclo:
            df_base = df_base[df_base["ciclo"].isin(ciclo)]

        if ruta:
            df_base = df_base[df_base["ruta"].isin(ruta)]
            
        # =========================
        # ESTADO ACTUAL DE FILTROS
        # =========================

        fise_sel_actual = st.session_state.get("filtro_fise_local", [])
        consignado_sel_actual = st.session_state.get("filtro_consignado_local", [])


        # =========================
        # CONSIGNADO DEPENDE DE (BASE + FISE)
        # =========================

        df_tmp_consignado = df_base.copy()

        if fise_sel_actual:
            df_tmp_consignado = df_tmp_consignado[
                df_tmp_consignado["cumplimiento_fise"].isin(fise_sel_actual)
            ]

        consignado_opciones = sorted(
            df_tmp_consignado["consignado"].dropna().unique()
        )


        # =========================
        # FISE DEPENDE DE (BASE + CONSIGNADO)
        # =========================

        df_tmp_fise = df_base.copy()

        if consignado_sel_actual:
            df_tmp_fise = df_tmp_fise[
                df_tmp_fise["consignado"].isin(consignado_sel_actual)
            ]

        fise_opciones = sorted(
            df_tmp_fise["cumplimiento_fise"].dropna().unique()
        )


        # =========================
        # LIMPIEZA DE SELECCIONES INVALIDAS
        # =========================

        st.session_state["filtro_fise_local"] = [
            x for x in st.session_state.get("filtro_fise_local", [])
            if x in fise_opciones
        ]

        st.session_state["filtro_consignado_local"] = [
            x for x in st.session_state.get("filtro_consignado_local", [])
            if x in consignado_opciones
        ]

        st.session_state["filtro_ruta_local"] = [
            x for x in st.session_state.get("filtro_ruta_local", [])
            if x in rutas_filtradas
        ]

        # ----------------------------
        # UI FILTROS (3 EN UNA FILA)
        # ----------------------------

        col1, col2, col3 = st.columns(3)

        with col1:
            ruta_seleccionada = st.multiselect(
                "Ruta",
                rutas_filtradas,
                key="filtro_ruta_local"
            )

        with col2:
            fise_sel = st.multiselect(
                "Cumplimiento FISE",
                fise_opciones,
                key="filtro_fise_local"
            )

        with col3:
            consignado_sel = st.multiselect(
                "Consignado",
                consignado_opciones,
                key="filtro_consignado_local"
            )

        # ----------------------------
        # APLICAR FILTROS
        # ----------------------------

        df = st.session_state.resumen_reparto.copy()

        df = aplicar_filtros_global(df)

        if ruta_seleccionada:
            df = df[df["ruta"].isin(ruta_seleccionada)]

        if fise_sel:
            df = df[df["cumplimiento_fise"].isin(fise_sel)]

        if consignado_sel:
            df = df[df["consignado"].isin(consignado_sel)]
               
        st.subheader("📊 Cumplimiento FISE")
        df_fise = df[
            ["lecturista","ciclo","ruta","total_suministros",
             "fise_obligatorio","fise_con_foto","fise_sin_foto","cumplimiento_fise"]
        ]

        st.dataframe(df_fise, use_container_width=True)
        
        st.subheader("📊 Cumplimiento 10% Fotos")

        df_10 = df[
            ["lecturista","ciclo","ruta","total_suministros",
            "fotos_tomadas","cant_min_foto","consignado"]
        ].copy()
        
        st.dataframe(df_10, use_container_width=True)
        
        archivo_excel = to_excel_unificado(
            df_mon,
            df_fise,
            df_10
        )

        st.download_button(
            "⬇ Descargar Reporte Completo",
            data=archivo_excel,
            file_name="seguimiento_reparto.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        