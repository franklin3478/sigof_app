import streamlit as st
import requests
import re
import pandas as pd
from openpyxl import Workbook
from io import BytesIO
from datetime import datetime

def ejecutar_selfie():

    mapa_unidades = {
        "Valle Mantaro": {"idempresa": "4", "iduunn": "83"},
        "Tarma": {"idempresa": "4", "iduunn": "79"},
        "Huancavelica": {"idempresa": "4", "iduunn": "78"},
        "Ayacucho": {"idempresa": "4", "iduunn": "76"},
        "Selva Central": {"idempresa": "4", "iduunn": "80"},
        "Huánuco": {"idempresa": "4", "iduunn": "82"},
        "Pasco": {"idempresa": "4", "iduunn": "81"},
        "Huancayo": {"idempresa": "4", "iduunn": "77"},
        "Tingo María": {"idempresa": "4", "iduunn": "84"},
    }

    observaciones_lista = [
        "CORRECTO",
        "Sin Epps",
        "Sin Fotocheck",
        "Sin Camisa",
        "Sin Chaleco",
        "Sin Camisa - Sin Chaleco - Sin Fotocheck.",
        "Sin Chaleco - Sin Fotocheck",
        "No se visualiza lecturador",
        "Selfie No corresponde al lecturador",
        "Fotografia incorrecta",
        "Sin Camisa - Sin Gorro",
        "Sin Chaleco - Sin Gorro",
        "Sin Fotocheck - Sin Camisa",
        "Sin Gorro"
    ]

    if "vista_selfie" not in st.session_state:
        st.session_state.vista_selfie = "login"

    if st.session_state.vista_selfie == "login":
        st.subheader("🔐 INGRESE USUARIO Y CONTRASEÑA")
        usuario = st.text_input("Usuario SIGOF")
        password = st.text_input("Contraseña SIGOF", type="password")
        if st.button("Ingresar"):
            if not usuario or not password:
                st.warning("Ingresa credenciales")
            else:
                st.session_state.usuario_sigof = usuario
                st.session_state.password_sigof = password
                st.session_state.vista_selfie = "sistema"
                st.rerun()

    if st.session_state.vista_selfie == "sistema":
        st.markdown(
            "<h2 style='text-align: center;'>VISUALIZACIÓN DE SELFIES</h2>",
            unsafe_allow_html=True
        )

        col1, col2 = st.columns(2)
        with col1:
            modulo = st.selectbox("🧩 Módulo", ["Lectura", "Reparto"])
        with col2:
            unidad = st.selectbox("🏢 Unidad de Negocio", list(mapa_unidades.keys()))

        st.session_state.modulo = modulo
        st.session_state.unidad = unidad

        # 🔹 limpiar datos al cambiar módulo o unidad
        if "modulo_anterior" not in st.session_state:
            st.session_state.modulo_anterior = modulo
            st.session_state.unidad_anterior = unidad

        if modulo != st.session_state.modulo_anterior or unidad != st.session_state.unidad_anterior:
            st.session_state.df_selfie = None
            st.session_state.modulo_anterior = modulo
            st.session_state.unidad_anterior = unidad

        df = st.session_state.get("df_selfie", None)

        def convertir_fecha(fecha_str):
            meses = {
                "January": "01", "February": "02", "March": "03", "April": "04",
                "May": "05", "June": "06", "July": "07", "August": "08",
                "September": "09", "October": "10", "November": "11", "December": "12"
            }
            match = re.match(r"(\d{1,2}) de ([a-zA-Z]+) de (\d{4}) en horas: (\d{2}:\d{2}:\d{2})", fecha_str)
            if match:
                dia, mes, anio, hora = match.groups()
                return f"{dia.zfill(2)}/{meses.get(mes,'00')}/{anio[-2:]}"
            return fecha_str.split(" ")[0]

        if df is not None:
            df["Fecha Selfie"] = df["Fecha Selfie"].apply(lambda x: convertir_fecha(x))

            def ordenar_fecha(f):
                try:
                    return datetime.strptime(f, "%d/%m/%y")
                except:
                    return datetime.min

            fechas_ordenadas = sorted(df["Fecha Selfie"].unique(), key=ordenar_fecha)
            opciones = fechas_ordenadas

            fechas_seleccionadas = st.multiselect(
                "📅 Fecha",
                opciones,
                placeholder="Seleccione una o más fechas"
            )

            if not fechas_seleccionadas:
                df_filtrado = df.copy().reset_index(drop=True)
            else:
                df_filtrado = df[df["Fecha Selfie"].isin(fechas_seleccionadas)].copy().reset_index(drop=True)
        else:
            st.selectbox("📅 Fecha", ["Primero procesa datos"], disabled=True)
            df_filtrado = None

        if st.button("📥 Obtener Selfies"):
            usuario = st.session_state.usuario_sigof
            password = st.session_state.password_sigof
            unidad = st.session_state.unidad
            modulo = st.session_state.modulo

            st.info("Conectando a SIGOF...")

            login_url = "http://sigof.distriluz.com.pe/plus/usuario/login"
            if modulo == "Lectura":
                url_modulo = "http://sigof.distriluz.com.pe/plus/ComlecOrdenlecturas/selfielecturista/"
                data_url = "http://sigof.distriluz.com.pe/plus/ComlecOrdenlecturas/ajax_mostar_mapa_selfie/"
            else:
                url_modulo = "http://sigof.distriluz.com.pe/plus/ComrepOrdenrepartos/"
                data_url = "http://sigof.distriluz.com.pe/plus/ComrepOrdenrepartos/ajax_mostar_mapa_selfie"

            headers = {
                "User-Agent": "Mozilla/5.0",
                "X-Requested-With": "XMLHttpRequest",
                "Referer": url_modulo
            }

            try:
                with requests.Session() as session:
                    session.post(login_url, data={
                        "data[Usuario][usuario]": usuario,
                        "data[Usuario][pass]": password
                    }, headers=headers)

                    session.get("http://sigof.distriluz.com.pe/plus/dashboard/modulos", headers=headers)

                    u = mapa_unidades[unidad]
                    session.post(
                        "http://sigof.distriluz.com.pe/plus/usuario/ajax_cambiar_sesion",
                        data={"idempresa": u["idempresa"], "iduunn": u["iduunn"]},
                        headers=headers
                    )

                    session.get("http://sigof.distriluz.com.pe/plus/dashboard/init", headers=headers)
                    session.get(url_modulo, headers=headers)
                    response = session.post(data_url, data={"_": "1"}, headers=headers)

                data = response.content.decode("utf-8", errors="ignore")
                if len(data.strip()) < 50:
                    st.error("❌ No se encontraron datos")
                    return

                data_cleaned = data.replace("\\/", "/")
                data_cleaned = re.sub(r"<\/?\w+.*?>", "", data_cleaned)
                data_cleaned = re.sub(r"\s+", " ", data_cleaned).strip()
                blocks = re.split(r"Ver detalle", data_cleaned)

                results = {}
                for block in blocks:
                    fecha = re.search(r"Fecha Selfie:\s*(\d{1,2} de [a-zA-Z]+ de \d{4} en horas: \d{2}:\d{2}:\d{2})", block)
                    persona = re.search(r"(Lecturista|Repartidor):\s*(.*?)\s*(Suministro Actual:|Fecha Selfie:|$)", block)
                    suministro = re.search(r"Suministro Actual:\s*(\d+)", block)
                    urls = re.findall(r'url":"(https?://[^"]+)"', block)

                    if fecha and persona and urls:
                        fecha_selfie = convertir_fecha(fecha.group(1).strip())
                        nombre_raw = re.sub(r"\s+", " ", persona.group(2)).strip()
                        nombre = bytes(nombre_raw, "utf-8").decode("unicode_escape")
                        key = (nombre, fecha_selfie)
                        if key not in results:
                            results[key] = []
                        for u in urls:
                            results[key].append((u, suministro.group(1) if suministro else ""))

                if results:
                    max_urls = max(len(v) for v in results.values())
                    columns = ["Fecha Selfie", "Lecturista"]
                    for i in range(max_urls):
                        columns.append(f"Url_foto {i+1}")
                        columns.append(f"Suministro_{i+1}")

                    data_list = []
                    for (nombre, fecha), urls in results.items():
                        row = [fecha, nombre]
                        for url, sumi in urls:
                            row.extend([url, sumi])
                        faltantes = max_urls - len(urls)
                        for _ in range(faltantes):
                            row.extend(["", ""])
                        data_list.append(row)

                    df = pd.DataFrame(data_list, columns=columns)

                    for fila_idx, (_, row) in enumerate(df.iterrows()):
                        for i in range(max_urls):
                            key = f"obs_{row['Lecturista']}_{i}_{row['Fecha Selfie']}"
                            st.session_state[key] = "CORRECTO"

                    st.session_state.df_selfie = df
                    st.rerun()
                else:
                    st.error("❌ Unidad válida pero sin selfies")

            except Exception as e:
                st.error(f"Error: {e}")

        if df_filtrado is not None:
            st.markdown("### 📋 Resultados")
            url_cols = [c for c in df.columns if "Url_foto" in c]
            max_imgs = len(url_cols)

            for _, row in df_filtrado.iterrows():
                col1, col2, col3 = st.columns([1, 2, 6])
                col1.write(row["Fecha Selfie"])
                col2.write(row["Lecturista"])
                cols_img = col3.columns(max_imgs)

                for i in range(max_imgs):
                    url = row[f"Url_foto {i+1}"]
                    sumi = row[f"Suministro_{i+1}"]
                    if url:
                        with cols_img[i]:
                            st.markdown(
                                f"""
                                <div style="text-align:center;">
                                    <img src="{url}" width="200"><br>
                                    {"<span>"+str(sumi)+"</span>" if modulo=="Reparto" else ""} 
                                </div>
                                """,
                                unsafe_allow_html=True
                            )
                            st.markdown("<div style='text-align:center;'>Observación</div>", unsafe_allow_html=True)
                            st.selectbox(
                                "",
                                observaciones_lista,
                                index=0,
                                key=f"obs_{row['Lecturista']}_{i}_{row['Fecha Selfie']}"
                            )
                    else:
                        cols_img[i].empty()

        # BLOQUE DE EXPORTACIÓN (FILTRADO)
        if df_filtrado is not None:
            output = BytesIO()
            wb = Workbook()
            ws = wb.active

            modulo = st.session_state.modulo
            col_persona = "Lecturista" if modulo == "Lectura" else "Repartidor"

            if col_persona not in df_filtrado.columns:
                otra_col = "Lecturista" if col_persona == "Repartidor" else "Repartidor"
                if otra_col in df_filtrado.columns:
                    df_filtrado[col_persona] = df_filtrado[otra_col]
                else:
                    df_filtrado[col_persona] = ""

            url_cols = [c for c in df_filtrado.columns if "Url_foto" in c]
            max_imgs = len(url_cols)

            def letra_columna(idx):
                letra = ""
                while idx > 0:
                    idx, rem = divmod(idx-1, 26)
                    letra = chr(65 + rem) + letra
                return letra

            headers = ["Fecha Selfie", col_persona] + \
                      [f"Url_foto_{i+1}" for i in range(max_imgs)] + \
                      [f"Imagen {i+1:02d}" for i in range(max_imgs)] + \
                      ["Observaciones"]
            ws.append(headers)

            fila_inicio = 2
            primera_columna_imagen = 3

            for fila_idx, (_, row) in enumerate(df_filtrado.iterrows(), start=fila_inicio):
                fila = [row["Fecha Selfie"], row[col_persona]]

                for i in range(max_imgs):
                    fila.append(row.get(f"Url_foto {i+1}", ""))

                for i in range(max_imgs):
                    col_letra = letra_columna(primera_columna_imagen + i)
                    if fila_idx == fila_inicio and i == 0:
                        fila.append(f"'=SI.ERROR(IMAGEN({col_letra}{fila_inicio};;3;250;180);\"\")")
                    else:
                        fila.append("")

                obs_dict = {}
                for i in range(max_imgs):
                    key = f"obs_{row[col_persona]}_{i}_{row['Fecha Selfie']}"
                    obs = st.session_state.get(key, "CORRECTO")
                    if obs != "CORRECTO":
                        if obs not in obs_dict:
                            obs_dict[obs] = []
                        obs_dict[obs].append(i+1)

                if not obs_dict:
                    fila.append("CORRECTO")
                else:
                    observaciones_lista_excel = []
                    for obs, imgs in obs_dict.items():
                        imgs_formateadas = [f"{img:02d}" for img in imgs]
                        if len(imgs_formateadas) == 1:
                            observaciones_lista_excel.append(f"imagen {imgs_formateadas[0]}: {obs}")
                        else:
                            if len(imgs_formateadas) == 2:
                                imgs_str = " y ".join(imgs_formateadas)
                            else:
                                imgs_str = ", ".join(imgs_formateadas[:-1]) + " y " + imgs_formateadas[-1]
                            observaciones_lista_excel.append(f"imagen {imgs_str}: {obs}")
                    fila.append("\n".join(observaciones_lista_excel))

                ws.append(fila)

            for r in range(2, ws.max_row + 1):
                ws.row_dimensions[r].height = 189

            for c in range(1, ws.max_column + 1):
                col_letra = letra_columna(c)
                ws.column_dimensions[col_letra].width = 25
                if "Url_foto" in ws.cell(row=1, column=c).value:
                    ws.column_dimensions[col_letra].hidden = True

            wb.save(output)
            st.download_button(
                "📥 Descargar Excel",
                data=output.getvalue(),
                file_name="Reporte_Selfie.xlsx"
            )